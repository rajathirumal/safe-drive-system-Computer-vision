import openpyxl
import argparse

ap = argparse.ArgumentParser()
ap.add_argument("-DayName", required=True,help="day name")
ap.add_argument("-month", required=True,help="month name")
ap.add_argument("-date", required=True,help="date")
ap.add_argument("-year", required=True,help="year")
ap.add_argument("-time", required=True,help="time")
ap.add_argument("-zone", required=True,help="zone")
ap.add_argument("-type", required=True,help="type")

args = vars(ap.parse_args())

wb = openpyxl.load_workbook('logs.xlsx')
for ws in wb.worksheets:
    print(ws.title)

ws = wb.worksheets[0]
'''
    python hello.py -DayName Sun -month Aug -date 18 -year 2019 -time 21:58:32 -zone GMT
'''
c = [args['date'], args['DayName'], args['month'], args['year'], args['time'], args['zone'],args['type']]

to_append = ws.cell(1,1).value

for i in range(len(c)):
    ws.cell(row=to_append+1, column=i+1).value = c[i]
to_append+=1    
ws.cell(1,1).value = to_append
wb.save("logs.xlsx")